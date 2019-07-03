# _*_ coding: utf-8 _*_
from . import _baobiao

import codecs
import pandas as pd
from flask import render_template, request, send_from_directory, abort, flash, redirect, send_file
from pymysql import err
from flask_login import login_required, current_user
import os
import re
from .form import BaobiaoForm, TianxieForm, QueryForm, PreviewForm
from pypinyin import lazy_pinyin
import pyexcel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
from win32com.client.gencache import EnsureDispatch
from datetime import datetime, date, timedelta
from win32com.client import Dispatch, constants
from .form import GenerateForm, excels
from .. import conn
from ..models import BaobiaoToSet
import pythoncom

basedir = os.path.abspath(os.path.dirname(__file__))
pardir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))


# 获取数据库中报表名
def get_baobiao_name():
    result = BaobiaoToSet.query.order_by(BaobiaoToSet.id).all()
    FILE_TO_SET = {}
    for i in range(len(result)):
        rs = str(result[i]).split(',')
        file = str(rs[0].strip('"').strip("'"))
        freq = str(rs[1].strip('"').strip("'"))
        FILE_TO_SET[str(i+1)] = file
    return FILE_TO_SET


def get_baobiao_freq():
    result = BaobiaoToSet.query.order_by(BaobiaoToSet.id).all()
    FREQ_OF_FILE = {}
    for i in range(len(result)):
        rs = str(result[i]).split(',')
        file = str(rs[0].strip('"').strip("'"))
        freq = str(rs[1].strip('"').strip("'"))
        FREQ_OF_FILE[file] = freq
    return FREQ_OF_FILE


@_baobiao.route('/split/')
@login_required
def split():
    form = BaobiaoForm()
    form.excels.choices = [(a.id, a.file) for a in BaobiaoToSet.query.all()]
    conn.ping(reconnect=True)
    cursor = conn.cursor()
    filetosetlist = request.values.getlist("excels")
    # print(filetosetlist)
    if filetosetlist == []:
        return render_template("baobiao.html", form=form)
    else:
        for file in filetosetlist:
            baobiao_split(cursor, file)
            conn.commit()
        flash('Baobiao(s) Successfully Splitted')
        conn.close()
        return render_template('baobiao.html', form=form)


def baobiao_split(cursor, file):
    FILE_TO_SET = get_baobiao_name()
    FREQ_OF_FILE = get_baobiao_freq()
    filetoset_chinese = FILE_TO_SET[file]
    filetoset = ''.join(lazy_pinyin(FILE_TO_SET[file])).lower()
    freq = FREQ_OF_FILE[filetoset_chinese]
    sql = 'select distinct position, content from ' + filetoset + ' where editable=True;'
    cursor.execute(sql)
    conn.commit()
    sqlresult = cursor.fetchall()
    userlist = []
    userset = {}
    lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
    lastmonth = lastmonthend.strftime("_%Y_%m")
    for i in range(len(sqlresult)):
        # 获取哪个格子
        position = sqlresult[i][0]
        # 获取计算式子
        ###################
        content_list = re.split('[-+*/()（）]', sqlresult[i][1].lstrip('|'))
        content_list = [s for s in content_list if s != '']
        # 存储分割后的数据进报表表
        content_store = ';'.join(content_list)
        sql = 'update ' + filetoset + ' set content_list="' + str(content_store) + '" where position="' + str(position) + '";'
        cursor.execute(sql)
        conn.commit()
        # 拆分到用户表
        for content in content_list:
            userandvalue = re.split(':|：', content)
            user = ''.join(lazy_pinyin(userandvalue[0]))
            if len(userandvalue) > 1:
                value = userandvalue[1]
            else:
                value = None
            if user not in userlist:
                userlist.append(user)
                userset[user] = []
            userset[user].append((position, value))
    for user in userlist:
        sql = 'create table if not exists ' + user + '(baobiao VARCHAR(100), position VARCHAR(100),' \
                ' content VARCHAR(500), value_last Double, value DOUBLE, submit_time VARCHAR(20), freq VARCHAR(5));'
        cursor.execute(sql)
        sql = 'delete from ' + user + ' where baobiao="' + filetoset_chinese + '";'
        cursor.execute(sql)
        try:
            sql = 'delete from ' + user + lastmonth + ' where baobiao="' + filetoset_chinese + '";'
            cursor.execute(sql)
        except:
            pass
        for i in range(len(userset[user])):
            # print(userset[user][i])
            position = userset[user][i][0]
            value = userset[user][i][1]
            sql = 'insert into ' + user + ' (baobiao, position, content, freq) values ("' + \
                  filetoset_chinese + '", "' + str(position) + '", "' + str(value) + '", "' + str(freq) + '");'
            cursor.execute(sql)
            try:
                sql = 'insert into ' + user + lastmonth + ' (baobiao, position, content, freq) values ("' + \
                      filetoset_chinese + '", "' + str(position) + '", "' + str(value) + '", "' + str(freq) + '");'
                cursor.execute(sql)
            except:
                pass


@_baobiao.route('/fill/', methods=['GET', 'POST'])
@login_required
def fill():
    form = TianxieForm()
    # 预览填写报表
    previewform = PreviewForm()
    previewform.excel.choices = [(a.id, a.file) for a in BaobiaoToSet.query.all()]

    username = current_user.username.lower()
    conn.ping(reconnect=True)
    cursor = conn.cursor()
    lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
    lastm = lastmonthend.strftime(("%m"))
    lastmonth = lastmonthend.strftime("_%Y_%m")
    try:
        if lastm in ["01", "02", "04", "05", "07", "08", "10", "11"]:
            sql = 'create table if not exists ' + username + lastmonth + \
                  ' select distinct * from ' + username + ' where freq="M";'
            cursor.execute(sql)
            try:
                # 上一期， Monthly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=1)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and position=s.position and content=s.content and freq="M");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            # 不考虑不同报表和格子位置，只要填写内容即只展现一次
            sql = 'select *, count(distinct content) from ' + username + lastmonth + ' where freq="M" group by content;'
            cursor.execute(sql)
            sqlresult = cursor.fetchall()
        elif lastm in ["03", "09"]:
            sql = 'create table if not exists ' + username + lastmonth + \
                  ' select distinct * from ' + username + ' where freq in ("M", "Q");'
            cursor.execute(sql)
            try:
                # 上一期, Monthly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=1)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and position=s.position and content=s.content and freq="M");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            try:
                # 上一期, Quarterly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=63)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and position=s.position and content=s.content and freq="Q");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            sql = 'select *, count(distinct content) from ' + username + lastmonth + ' where freq in ("M", "Q") group by content;'
            cursor.execute(sql)
            sqlresult = cursor.fetchall()
        elif lastm in ["06", "12"]:
            sql = 'create table if not exists ' + username + lastmonth + \
                  ' select distinct * from ' + username + ';'
            cursor.execute(sql)
            try:
                # 上一期, Monthly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=1)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and position=s.position and content=s.content and freq="M");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            try:
                # 上一期, Quarterly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=63)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and position=s.position and content=s.content and freq="Q");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            try:
                # 上一期, Half year
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=153)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and position=s.position and content=s.content and freq="H");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            sql = 'select *, count(distinct content) from ' + username + lastmonth + ' group by content;'
            cursor.execute(sql)
            sqlresult = cursor.fetchall()
    except:
        sqlresult = None
    if request.method == 'GET' and sqlresult is not None:
        conn.close()
        return render_template("baobiao_tianxie.html", form=form, previewform=previewform, sqlresult=sqlresult)
    elif request.method == 'GET' and sqlresult is None:
        conn.close()
        return render_template("baobiao_tianxie.html", form=form, previewform=previewform)
    elif request.method == 'POST':
        # 预览填写报表
        if previewform.excel.data is not None:
            FILE_TO_SET = get_baobiao_name()
            baobiao = FILE_TO_SET[str(previewform.excel.data)]
            print(baobiao)
            filedir = os.path.join(pardir, 'Files', 'upload')
            destdir = os.path.join(pardir, 'templates')
            # pyexcel.save_as(file_name=filedir + '/' + baobiao + '.xlsx',
            #                 dest_file_name=destdir + '/preview.handsontable.html')
            xd = pd.ExcelFile(filedir + '/' + baobiao + '.xlsx')
            pd.set_option('display.max_colwidth', 1000)
            df = xd.parse()
            with codecs.open(destdir + '/preview.html', 'w', encoding='utf-8') as html_file:
                html_file.write(df.to_html(header=True, index=True, na_rep=''))
            return render_template("preview.html")

        # 填写内容
        tianxie = request.form.getlist("values")
        try:
            for i in range(len(tianxie)):
                # baobiao = sqlresult[i][0]
                # gezi = sqlresult[i][1]
                content = sqlresult[i][2]
                value = str(tianxie[i])
                submit_time = datetime.today().strftime('%y/%m/%d %H:%M')
                sql = 'update ' + username + lastmonth + ' set value=' + value + ' where content="' + content + '";'
                sql2 = 'update ' + username + lastmonth + ' set submit_time="' + submit_time + '" where content="' + content + '";'
                if value != '':
                    print(sql)
                    print(sql2)
                    cursor.execute(sql)
                    cursor.execute(sql2)
            conn.commit()
            conn.close()
            flash("数据提交成功")
            return redirect('/baobiao/fill/')
        except:
            flash("有个格子不是数字格式，只能填写数字格式")
            return redirect('/baobiao/fill/')
        finally:
            return redirect('/baobiao/fill/')


@_baobiao.route('/generate')
@login_required
def generate():
    form = GenerateForm()
    form.excels.choices = [(a.id, a.file) for a in BaobiaoToSet.query.all()]
    FILE_TO_SET = get_baobiao_name()
    generatelist = request.values.getlist('excels')
    filedir = os.path.join(pardir, 'Files', 'generate')
    if not os.path.exists(filedir):
        os.mkdir(filedir)
    if generatelist == []:
        return render_template('generate.html', form=form)
    else:
        lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
        lastmonth = lastmonthend.strftime("%Y_%m_%d")
        print(lastmonth)
        generatedate = lastmonth
        allcomplete = True
        for generatefile in generatelist:
            filetogenerate_chinese = FILE_TO_SET[generatefile]
            alert = generateFile(filetogenerate_chinese, generatedate)
            if len(alert) != 0:
                allcomplete = False
                alertmsg = filetogenerate_chinese + ': ' + ','.join(alert)
                flash('以下用户还未完成对应报表：' + alertmsg)
        if allcomplete:
            flash('报表生成成功')
        else:
            flash('报表生成成功但数据还未完整')
        return render_template('generate.html', form=form)


def generateFile(filetogenerate_chinese, generatedate):
    pythoncom.CoInitialize()
    conn.ping(reconnect=True)
    cursor = conn.cursor()
    filetogenerate = ''.join(lazy_pinyin(filetogenerate_chinese))
    tablenamenew = filetogenerate + '_' + generatedate
    userlist = []
    # 获取上个月日期
    lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
    lastmonth = lastmonthend.strftime("_%Y_%m")
    # 创建新表
    sql = 'drop table if exists ' + tablenamenew
    cursor.execute(sql)
    sql = 'create table ' + tablenamenew + \
          '(tablename VARCHAR(100), position VARCHAR(100), content VARCHAR(100),' \
          ' editable Boolean, content_formula VARCHAR(500), content_list VARCHAR(500), primary key (position));'
    cursor.execute(sql)
    conn.commit()
    try:
        sql = 'insert into ' + tablenamenew + ' (tablename, position, content, editable, content_formula, content_list) ' \
              'select tablename, position, content, editable, content, content_list from ' + filetogenerate + ';'
        cursor.execute(sql)
        conn.commit()
    except:
        print('Update Table Failure')
    finally:
        pass

    # 从模板拿需要填写的格子
    sql = 'select distinct position, content, content_list from ' + filetogenerate + ' where editable=True;'
    cursor.execute(sql)
    # conn.commit()
    sqlresult = cursor.fetchall()
    # 用来提示哪些用户还未填写此张报表
    alertset = set()
    # 按每一个格子循环去用户表中取数
    for i in range(len(sqlresult)):
        # 获取哪个格子
        position = sqlresult[i][0]
        formula = sqlresult[i][1]
        # 获取用户和内容
        content_list = sqlresult[i][2].split(';')
        uservalue_list = []
        for content in content_list:
            userandvalue = re.split(':|：', content)
            user = ''.join(lazy_pinyin(userandvalue[0]))
            if user not in userlist:
                userlist.append(user)
            if len(userandvalue) > 1:
                valuecontent = userandvalue[1]
            else:
                valuecontent = None
            try:
                sql = 'select value from ' + user + lastmonth + ' where baobiao="' + filetogenerate_chinese + \
                      '" and position="' + str(position) + '" and content="' + valuecontent + '";'
                cursor.execute(sql)
                value = cursor.fetchone()[0]
                uservalue_list.append(value)
            except err.DatabaseError:
                value = None
                uservalue_list.append(value)
            else:
                pass
            # 代入运算式
            # 需要按式子中的顺序查找对应的值，用uservalue_list代替content_list中的值并代入formula
            # print(userandvalue)
            # print(value)
            if value is None:
                alertset.add(user)
                formula = formula.replace(content, str(0))
            else:
                formula = formula.replace(content, str(value))
        formula = formula.replace("（", "(")
        formula = formula.replace("）", ")")
        # print(formula)
        try:
            positionresult = round(eval(formula.lstrip("|")), 2)
        except:
            positionresult = "=NA()" ##计算报错就在excel中填NA
        sql = 'update ' + filetogenerate + '_' + generatedate + ' set content="' + str(positionresult) + \
              '" where position="' + str(position) + '";'
        cursor.execute(sql)
        conn.commit()
    ######################
    # 生成excel
    # 计算行数列数
    wb = load_workbook(pardir + '/Files/upload/' + filetogenerate_chinese + '.xlsx')
    sh = wb.active
    sql = 'select distinct position, content from ' + filetogenerate + '_' + generatedate + ' where editable=TRUE;'
    cursor.execute(sql)
    conn.commit()
    sqlresult = cursor.fetchall()
    for x in sqlresult:
        try:
            sh[x[0]] = float(x[1])
            sh[x[0]].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        except:
            sh[x[0]] = x[1]
    # 把带公式计算的格子填入公式，自动计算
    sql = 'select distinct position, content from ' + filetogenerate + ' where content like "=%";'
    cursor.execute(sql)
    conn.commit()
    sqlresult = cursor.fetchall()
    for x in sqlresult:
        sh[x[0]] = str(x[1])
        sh[x[0]].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    filedir = os.path.join(pardir, 'Files', 'Generate', filetogenerate_chinese)
    if not os.path.exists(filedir):
        os.mkdir(filedir)
    # 保存带公式的xlsx
    wb.save(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx')
    # 去除公式只保存数值,需要先excel程序打开再保存一下，然后用openpyxl只保留数值，最后再存为html用于预览
    #### https://www.cnblogs.com/vhills/p/8327918.html
    xlApp = EnsureDispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx')
    xlBook.Save()
    xlBook.Close()
    xlApp.Quit()
    wb = load_workbook(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx', data_only=True)
    # 删除除了第一个sheet外的sheet
    # sheetnames = wb.get_sheet_names()
    # for i in range(1, len(sheetnames)):
    #     sheet = wb.get_sheet_by_name(sheetnames[i])
    #     wb.remove_sheet(sheet)
    wb.save(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx')
    return alertset


@_baobiao.route('/query/', methods=['GET', 'POST'])
@login_required
def query():
    pythoncom.CoInitialize()
    form = QueryForm()
    form.excel.choices = [(a.id, a.file) for a in BaobiaoToSet.query.all()]
    if request.method == 'GET':
        return render_template("baobiao_query.html", form=form)
    else:
        FILE_TO_SET = get_baobiao_name()
        baobiao = FILE_TO_SET[str(form.excel.data)]
        baobiao = "".join(lazy_pinyin(baobiao))
        generatedate = request.values.get('generatedate')
        generatedate = generatedate.replace('-', '_')
        lastdate = request.values.get('lastdate')
        lastdate = lastdate.replace('-', '_')
        baobiao_generate = baobiao + '_' + generatedate
        baobiao_last = baobiao + '_' + lastdate
        filedir = os.path.join(pardir, 'Files', 'Generate', baobiao)
        destdir = os.path.join(pardir, 'templates')
        ######

        result = baobiao_compare(baobiao, generatedate, lastdate)
        return render_template("baobiao_query_result.html", form=form, result=result, baobiao=baobiao)


def baobiao_compare(baobiao, generatedate, lastdate):
    baobiao_generate = "".join(lazy_pinyin(baobiao + '_' + generatedate.replace('-', '_')))
    baobiao_last = "".join(lazy_pinyin(baobiao + '_' + lastdate.replace('-', '_')))
    conn.ping(reconnect=True)
    cursor = conn.cursor()
    # this term baobiao
    sql = 'select distinct position, content, content_formula from ' + baobiao_generate + ' where editable=True;'
    cursor.execute(sql)
    conn.commit()
    sqlresult_generate = cursor.fetchall()
    result_generate = dict((x, [z, y]) for x, y, z in sqlresult_generate)
    # print(result_generate)
    # last term baobiao
    sql = 'select distinct position, content from ' + baobiao_last + ' where editable=True;'
    cursor.execute(sql)
    conn.commit()
    sqlresult_last = cursor.fetchall()
    result_last= dict((x, y) for x, y in sqlresult_last)
    changedict = {}
    for key in result_generate:
        try:
            value_explain = result_generate.get(key)[0].split('|')
            value_explain = "，".join(value_explain).lstrip('，')
            value_generate = result_generate.get(key)[1]
            value_last = result_last.get(key)
            # print(value_generate)
            print(value_last)
            if float(value_last) != 0:
                pctchange = float(value_generate) / float(value_last) - 1
                show_result = str(round(pctchange * 100, 2)) + '%'
            else:
                pctchange = None
                show_result = 'Cannot Compare Percentage Change'
            changedict[key] = [value_explain, value_last, value_generate, show_result, pctchange]
            print(pctchange)
        except KeyError:
            print('No key in last table')
            value_generate = result_generate.get(key)
            value_last = None
            pctchange = None
            changedict[key] = [value_explain, value_last, value_generate, None, pctchange]
        finally:
            pass
    # print(changedict)
    return changedict


@_baobiao.route('/query/generate', methods=['GET', 'POST'])
@login_required
def show_generate_baobiao():
    # return send_file(os.path.join(pardir, "templates", "query_report", "query_report.html"))
    # return render_template("query_report.html")
    return render_template("query_report.handsontable.html")


@_baobiao.route('/query/last', methods=['GET', 'POST'])
@login_required
def show_last_baobiao():
    return render_template("last_report.handsontable.html")
